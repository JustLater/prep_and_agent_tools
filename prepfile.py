from pydantic import BaseModel, Field
from typing import Optional, Tuple, List, Dict, Any
import re
import base64
from pathlib import Path
import pandas as pd
import numpy as np
from datetime import datetime
import openpyxl
import xlrd
import os
import traceback
import json
import math
import pymystem3
from pymystem3 import Mystem
import string
from collections import Counter
from openai import OpenAI, DefaultHttpxClient
import torch
from transformers import T5Tokenizer, T5ForConditionalGeneration

class DataStorage:
    def __init__(self, name: str, file_id: str):
        # Имя и id, совпадающее с id файла
        self.name = name
        self.id = file_id
        # Датафрейм с оригинальной таблицей
        self.orig_df = pd.DataFrame()
        # Очищенный от "ненужных" столбцов датафрейм
        self.cleaned_df = pd.DataFrame()
        # Датафрейм с целевыми столбцами
        self.target_df = pd.DataFrame()
        # Список названий столбцов, где мало ответов
        self.few_answers_list = []
        # Датафрейм с уникальными значениями по столбцам
        self.unique_values_df = pd.DataFrame()
        # Словарь с количеством уникальных значений на столбец
        self.unique_dict = {}

        # Датафрейм с тестовыми столбцами
        self.test_df = pd.DataFrame()
        # Датафрейм с nlp столбцами
        self.nlp_df = pd.DataFrame()
        # Датафрейм со столбцами, где мало ответов
        self.late_nlp_df = pd.DataFrame()

        # Закодированный тестовый датафрейм
        self.test_ohe_df = pd.DataFrame()
        # Лемматизированный nlp датафрейм
        # self.nlp_lemmed_df = pd.DataFrame()
        # Закодированный словарь (выбранные слова и 0 или 1 в строках)
        self.nlp_ohe_dict = {}
        # Векторизованный нлп датафрейм
        self.nlp_vec_df = pd.DataFrame()
        # Датафрейм с длинными отзывами
        self.nlp_large_df = pd.DataFrame()

    def preprocess(self, preprocessor: "Preprocessor") -> None:
        """
        Метод-обёртка для запуска всего пайплайна препроцессинга.
        Делегирует логику классу Preprocessor.
        """
        preprocessor.preprocess_tables(self)

class Preprocessor:
    """Класс для операций предоработки"""

    def preprocess_tables(self, storage: DataStorage) -> None:
        """
        Препроцесс таблицы

        Args:
            storage: аргумент класса, хранящий информацию о таблице
        Returns:
            None: заполняет объект класса данными
        """
        # storage.orig_df = self.parse_merged_rows(input_path)
        # storage.orig_df = self.find_columns(input_path)
        storage.few_answers_list = self.find_no_answer(storage.orig_df)
        print("DEBUG: storage.few_answers_list заполнен")
        storage.unique_values_df = self.build_uvd(storage.orig_df)
        print("DEBUG: storage.unique_values_df заполнен")
        storage.unique_dict = self.count_unique_val(storage.unique_values_df)
        print("DEBUG: storage.unique_dict заполнен")
        storage.test_df, storage.nlp_df, storage.late_nlp_df = self.parse_df(
            storage.orig_df,
            storage.unique_dict,
            storage.few_answers_list,
            storage.target_df,
        )
        print("DEBUG: storage.test_df, nlp_df, late_nlp_df заполнены")
        storage.nlp_df, storage.nlp_large_df = self.find_large_answers(storage.nlp_df)
        print("DEBUG: storage.nlp_df, nlp_large_df заполнены")
        storage.test_ohe_df = self.encodeDf(
            storage.test_df, storage.unique_values_df, storage.target_df
        )
        print("DEBUG: storage.test_ohe_df заполнены")
        # опечатки
        # storage.nlp_lemmed_df, storage.nlp_ohe_dict = self.lemmatize_and_ohe_df(storage.nlp_df, mystem, stop_words)
        

    def parse_merged_rows(self, file_path: str) -> pd.DataFrame:
        """
        Функция для парсинга Excel-файлов с объединенными ячейками.

        Обрабатывает файлы Excel, корректно извлекая данные из объединенных строк,
        преобразуя типы данных.

        Args:
            file_path: Путь к исходному Excel-файлу для обработки.

        Returns:
            DataFrame с обработанными данными или пустой DataFrame в случае ошибки.
        """
        try:
            file_ext = os.path.splitext(file_path)[1].lower()

            if file_ext == ".csv":
                df = pd.read_csv(file_path)
                return df

            if file_ext == ".xls":
                book = xlrd.open_workbook(file_path)
                sheet = book.sheet_by_index(0)
                use_openpyxl = False
            elif file_ext == ".xlsx":
                book = openpyxl.load_workbook(file_path, data_only=True)
                sheet = book.active
                use_openpyxl = True
            else:
                raise ValueError(
                    "Неподдерживаемый формат файла. Используйте XLS, XLSX или CSV."
                )

            data = []
            headers = []
            current_row = None
            is_header = True

            if use_openpyxl:
                max_row = sheet.max_row
                max_col = sheet.max_column
            else:
                max_row = sheet.nrows
                max_col = sheet.ncols

            if max_row == 0 or max_col == 0:
                return pd.DataFrame()

            for row_idx in range(max_row):
                if use_openpyxl:
                    first_col_value = sheet.cell(row=row_idx + 1, column=1).value
                else:
                    first_col_value = sheet.cell_value(row_idx, 0)

                if first_col_value and str(first_col_value).strip():
                    if current_row:
                        data.append(current_row)

                    current_row = []
                    for col_idx in range(max_col):
                        if use_openpyxl:
                            cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)
                            cell_value = cell.value
                            ctype = (
                                2
                                if isinstance(cell_value, (int, float))
                                else (3 if isinstance(cell_value, datetime) else 1)
                            )
                        else:
                            cell = sheet.cell(row_idx, col_idx)
                            cell_value = cell.value
                            ctype = cell.ctype

                        if is_header:
                            headers.append(
                                str(cell_value)
                                if cell_value is not None
                                else f"Column_{col_idx+1}"
                            )
                        else:
                            if ctype == 3 or isinstance(cell_value, datetime):
                                if use_openpyxl:
                                    dt = cell_value
                                else:
                                    dt = xlrd.xldate_as_datetime(
                                        cell_value, book.datemode
                                    )
                                current_row.append(dt)
                            else:
                                try:
                                    if isinstance(cell_value, (int, float)):
                                        if cell_value == int(cell_value):
                                            current_row.append(int(cell_value))
                                        else:
                                            current_row.append(float(cell_value))
                                    else:
                                        current_row.append(
                                            np.nan
                                            if not cell_value
                                            else str(cell_value)
                                        )
                                except (ValueError, TypeError):
                                    current_row.append(
                                        np.nan if not cell_value else str(cell_value)
                                    )

                    if is_header:
                        is_header = False
                        current_row = None
                        continue

                elif current_row is not None:
                    for col_idx in range(max_col):
                        if use_openpyxl:
                            cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)
                            cell_value = cell.value
                            ctype = (
                                2
                                if isinstance(cell_value, (int, float))
                                else (3 if isinstance(cell_value, datetime) else 1)
                            )
                        else:
                            cell = sheet.cell(row_idx, col_idx)
                            cell_value = cell.value
                            ctype = cell.ctype

                        if cell_value and str(cell_value).strip():
                            if ctype == 3 or isinstance(cell_value, datetime):
                                if use_openpyxl:
                                    dt = cell_value
                                else:
                                    dt = xlrd.xldate_as_datetime(
                                        cell_value, book.datemode
                                    )
                                cell_value = dt
                            else:
                                try:
                                    if isinstance(cell_value, (int, float)):
                                        if cell_value == int(cell_value):
                                            cell_value = int(cell_value)
                                        else:
                                            cell_value = float(cell_value)
                                except (ValueError, TypeError):
                                    cell_value = str(cell_value)

                            if col_idx < len(current_row):
                                if (
                                    current_row[col_idx]
                                    and pd.notna(current_row[col_idx])
                                    and current_row[col_idx] != ""
                                ):
                                    current_row[col_idx] = (
                                        f"{current_row[col_idx]}\n{cell_value}"
                                    )
                                else:
                                    current_row[col_idx] = cell_value
                else:
                    if not is_header and current_row is None:
                        current_row = []
                        for col_idx in range(max_col):
                            if use_openpyxl:
                                cell = sheet.cell(row=row_idx + 1, column=col_idx + 1)
                                cell_value = cell.value
                                ctype = (
                                    2
                                    if isinstance(cell_value, (int, float))
                                    else (3 if isinstance(cell_value, datetime) else 1)
                                )
                            else:
                                cell = sheet.cell(row_idx, col_idx)
                                cell_value = cell.value
                                ctype = cell.ctype

                            if ctype == 3 or isinstance(cell_value, datetime):
                                if use_openpyxl:
                                    dt = cell_value
                                else:
                                    dt = xlrd.xldate_as_datetime(
                                        cell_value, book.datemode
                                    )
                                current_row.append(dt)
                            else:
                                try:
                                    if isinstance(cell_value, (int, float)):
                                        if cell_value == int(cell_value):
                                            current_row.append(int(cell_value))
                                        else:
                                            current_row.append(float(cell_value))
                                    else:
                                        current_row.append(
                                            np.nan
                                            if not cell_value
                                            else str(cell_value)
                                        )
                                except (ValueError, TypeError):
                                    current_row.append(
                                        np.nan if not cell_value else str(cell_value)
                                    )

            if current_row:
                data.append(current_row)

            if not data:
                return pd.DataFrame()

            if len(data[0]) != len(headers):
                headers = headers[: len(data[0])]

            df = pd.DataFrame(data, columns=headers)

            for col in df.columns:
                if any(isinstance(x, datetime) for x in df[col] if pd.notna(x)):
                    df[col] = pd.to_datetime(df[col], errors="coerce")
                else:
                    try:
                        numeric_series = pd.to_numeric(df[col], errors="coerce")
                        if numeric_series.notna().any():
                            if all(
                                numeric_series.dropna()
                                == numeric_series.dropna().astype(int)
                            ):
                                df[col] = numeric_series.astype("Int64")
                            else:
                                df[col] = numeric_series.astype(float)
                    except:
                        pass

            return df

        except Exception as e:
            print(f"Ошибка в parse_merged_rows: {e}")
            return pd.DataFrame()

    def find_no_answer(
        self, input_df: pd.DataFrame, no_answer_per: int = 30
    ) -> List[str]:
        """
        Функция для поиска столбцов с большим количеством пропущенных значений.

        Анализирует DataFrame и возвращает список столбцов, в которых количество
        пропущенных/пустых значений превышает заданный порог.

        Args:
            input_df: Исходный DataFrame для анализа.
            no_answer_per: Минимальное количество пропущенных значений (в процентах)
                          для включения столбца в результат. По умолчанию 30.

        Returns:
            List[str]: Список названий столбцов, содержащих указанное количество
                      или более пропущенных значений.
        """
        curr_columns = input_df.columns.copy()
        result = []
        answer_count = len(input_df)
        no_answer_count = math.ceil(answer_count * (no_answer_per / 100))
        for col in curr_columns:
            no_answer = 0
            for value in input_df[col].astype(str):
                if (
                    pd.isna(value)
                    or value == "nan"
                    or value == "None"
                    or value == ""
                    or value == "-"
                    or value == "_"
                ):
                    no_answer += 1
            if no_answer >= no_answer_count:
                result.append(col)
        return result

    def build_uvd(self, input_df: pd.DataFrame) -> pd.DataFrame:
        """
        Функция для построения словаря уникальных значений из DataFrame

        Создает DataFrame, содержащий все уникальные значения из каждого столбца исходных данных.

        Args:
            input_df: Исходный DataFrame для обработки

        Returns:
            pd.DataFrame: новый DataFrame, где:
                каждый столбец соответствует столбцу исходного DataFrame
                строки содержат уникальные значения из исходного столбца
        """
        unique_values_dict = {}

        for column in input_df:
            all_lines = []

            for cell in input_df[column]:
                if pd.isna(cell):
                    continue
                if isinstance(cell, str):
                    lines_in_cell = [
                        line.strip() for line in cell.split("\n") if line.strip()
                    ]
                    all_lines.extend(lines_in_cell)
                else:
                    all_lines.append(str(cell))

            unique_lines = sorted(set(filter(None, all_lines)), key=lambda x: str(x))
            unique_values_dict[column] = unique_lines

        max_len = (
            max(len(v) for v in unique_values_dict.values())
            if unique_values_dict
            else 0
        )
        padded_dict = {
            k: v + [np.nan] * (max_len - len(v)) for k, v in unique_values_dict.items()
        }

        return pd.DataFrame(padded_dict)

    def count_unique_val(self, input_df: pd.DataFrame) -> Dict[str, int]:
        """
        Функция для подсчета количества уникальных значений в каждом столбце DataFrame.

        Args:
            input_df: Исходный DataFrame для анализа.

        Returns:
            Dict[str, int]: Словарь, где:
                - Ключи: названия столбцов исходного DataFrame
                - Значения: количество уникальных значений в соответствующем столбце
        """
        unique_dict = {}
        for col in input_df.columns:
            unique_val = input_df[col].dropna().unique()
            unique_dict[col] = len(unique_val)
        return unique_dict

    def parse_df(
        self,
        input_df: pd.DataFrame,
        input_dict: Dict[str, int],
        late_cols: List[str],
        tar_cols: pd.DataFrame,
    ) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        Функция для разделения DataFrame на несколько частей по заданным критериям.

        Args:
            input_df: Исходный DataFrame для обработки.
            input_dict: Словарь с информацией о количестве уникальных значений в столбцах.
            late_cols: Список столбцов для выделения в отдельный DataFrame.
            tar_cols: DataFrame, который будет добавлен ко всем результирующим DataFrame (целевые столбцы).

        Returns:
            Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]: Кортеж из трех DataFrame:
                - test_df: Содержит столбцы с малым количеством уникальных значений (<20)
                - nlp_df: Содержит оставшиеся столбцы
                - late_df: Содержит столбцы, указанные в late_cols
        """
        df_copy = input_df.copy()
        df_copy.drop(columns=tar_cols, inplace=True)

        late_df = df_copy[late_cols].copy()
        df_copy.drop(columns=late_cols, inplace=True)

        test_cols = [
            col for col in input_dict if input_dict[col] < 20 and col in df_copy.columns
        ]
        test_df = df_copy[test_cols].copy()
        df_copy.drop(columns=test_cols, inplace=True)

        nlp_df = df_copy.copy()

        for df in [test_df]:
            for col in tar_cols.columns:
                df[col] = tar_cols[col].copy()

        return test_df, nlp_df, late_df

    def find_large_answers(
        self, input_df: pd.DataFrame, threshold: int = 10
    ) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Выделяет большие отзывы в отдельный датафрейм на основе заданного порога.
        """
        df = input_df.copy()
        mask = pd.DataFrame(False, index=df.index, columns=df.columns)

        for col in df.columns:
            col_str = df[col].astype(str)

            non_empty_df = col_str.replace("", np.nan).replace("nan", np.nan).dropna()

            if len(non_empty_df) == 0:
                continue

            avg_len = non_empty_df.str.split().str.len().mean()

            if pd.isna(avg_len) or avg_len == 0:
                continue

            mask[col] = (
                df[col].notna()
                & (col_str != "")
                & (col_str != "nan")
                & (col_str.str.len() > threshold * avg_len)
            )

        large_answ_df = df[mask].copy()
        df[mask] = ""

        return df, large_answ_df

    def encodeDf(
        self, input_df: pd.DataFrame, unique_df: pd.DataFrame, tar_df: pd.DataFrame
    ) -> pd.DataFrame:
        """
        Функция для кодирования категориальных признаков в DataFrame методом бинаризации.

        Выполняет one-hot encoding для мультизначений, где значения разделены через '\\n'.

        Args:
            input_df: Исходный DataFrame для обработки.
            unique_df: DataFrame, содержащий уникальные значения для каждого столбца, подлежащего кодированию.
            tar_df: DataFrame с целевыми переменными, которые будут добавлены в результат без изменений.

        Returns:
            DataFrame, где:
                - категориальные признаки заменены бинарными столбцами (по одному для каждого уникального значения)
                - для мультизначений устанавливается 1, если значение присутствует в строке
                - целевые переменные добавляются в результат без изменений
        """
        result_df = (
            input_df.drop(columns=tar_df.columns, errors="ignore").dropna().copy()
        )
        columns_to_encode = [
            col
            for col in unique_df.columns
            if col in input_df.columns and col not in tar_df.columns
        ]

        for column in columns_to_encode:
            unique_values = unique_df[column].dropna().unique()
            temp_df = pd.DataFrame(index=input_df.index)

            for value in unique_values:
                if value is not None:
                    new_col_name = f"{column}_{value}"
                    temp_df[new_col_name] = (
                        input_df[column]
                        .astype(str)
                        .apply(
                            lambda x: (
                                1
                                if str(value).strip()
                                in [v.strip() for v in str(x).split("\n")]
                                else 0
                            )
                        )
                    )
            result_df = result_df.drop(column, axis=1)
            result_df = pd.concat([result_df, temp_df], axis=1)

        cols_to_drop = [
            col
            for col in result_df.columns
            if any(suffix in col.lower() for suffix in ["_nan", "_none"])
        ]
        result_df = result_df.drop(columns=cols_to_drop, errors="ignore")

        result_df = pd.concat([result_df, tar_df], axis=1)
        return result_df

    def vectorize_df(self, input_df: pd.DataFrame, vec_model: Any, embedding_model_name: str) -> pd.DataFrame:
        """
        Функция для векторизации текстовых столбцов датафрейма с использованием модели эмбеддингов.

        Заменяет тексты на их векторные представления, для пустых значений использует нулевые векторы.

        Args:
            input_df: DataFrame с текстовыми данными.
            vec_model: модель векторизации.
            embedding_model_name: Имя модели для эмбеддингов.

        Returns:
            DataFrame: Копия датафрейма с векторизованными столбцами.
        """
        lemmed_input_not = input_df.copy()
        columns_to_vectorize = input_df.columns

        test_response = vec_model.embeddings.create(
            model=embedding_model_name, input=["test"]
        )
        dim = len(test_response.data[0].embedding)
        zero_vector = np.zeros(dim).tolist()

        for column in columns_to_vectorize:
            texts = lemmed_input_not[column].astype(str)

            mask = (texts != "") & (texts != "nan")

            valid_texts = texts[mask].tolist()
            all_embeddings = []

            if valid_texts:
                response = vec_model.embeddings.create(
                    model=embedding_model_name, input=valid_texts
                )
                all_embeddings = [item.embedding for item in response.data]

            embeddings_list = []
            embedding_idx = 0

            for is_valid in mask:
                if is_valid and embedding_idx < len(all_embeddings):
                    emb = all_embeddings[embedding_idx]
                    embedding_idx += 1
                    embeddings_list.append(
                        np.nan_to_num(emb, nan=0.0, posinf=1.0, neginf=-1.0).tolist()
                    )
                else:
                    embeddings_list.append(zero_vector)

            lemmed_input_not[column] = embeddings_list

        return lemmed_input_not

    def lemmatize_text(self, text: str, mystem: Mystem, stop_words: List[str]) -> str:
        """
        Функция для лемматизации текста с помощью pymystem3 с удалением стоп-слов.

        Args:
            text: Текст для лемматизации.
            mystem: Лемматизатор pymystem3.
            stop_words: Список стоп-слов для фильтрации.

        Returns:
            str: Лемматизированный текст без стоп-слов.
        """

        if not text or pd.isna(text):
            return ""

        lemmas = mystem.lemmatize(text.lower())
        return " ".join(
            [
                lemma.strip()
                for lemma in lemmas
                if lemma.strip() and lemma.strip() not in stop_words
            ]
        )

    def clean_text(self, text: str) -> str:
        """
        Функция для очистки текста от лишних символов и привидения к единому формату.

        Args:
            text: Текст для очистки.

        Returns:
            str: Очищенный текст.
        """
        if pd.isna(text):
            return ""

        text = str(text).strip().lower()
        text = re.sub(r"ё", "е", text)
        text = re.sub(r"/", " или ", text)
        text = re.sub(r"\[(.*?)\]", "", text)
        text = re.sub(r"\s+", " ", text)
        text = re.sub(r"\w+…|…", "", text)
        text = re.sub(r"(?<=\w)-(?=\w)", " ", text)
        text = re.sub(f"[{re.escape(string.punctuation)}]", "", text)
        return text

    def lemmatize_and_ohe_df(
        self,
        input_df: pd.DataFrame,
        mystem: Mystem,
        stop_words: List[str],
        min_per: int = 3,
        max_per: int = 90,
    ) -> Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]:
        """
        Функция для обработки DataFrame: очищает текст, лемматизирует
        и создает бинарное представление отдельных слов.

        Бинарное представление слов - адаптация one-hot-encoder.

        Args:
            input_df: Входной DataFrame с текстовыми данными.
            mystem: Лемматизатор pymystem3.
            stop_words: Список стоп-слов для фильтрации.
            min_per: Минимальная частота слова для включения в представление (в процентах). По умолчанию 3.
            max_per: Максимальная частота слова для включения в представление (в процентах). По умолчанию 90.

        Returns:
            Tuple[pd.DataFrame, Dict[str, pd.DataFrame]]: Кортеж из:
                - лемматизированный DataFrame
                - словарь с бинарным представлением слов по каждому столбцу
        """

        lemmed_df = input_df.copy()

        min_count = math.ceil(len(lemmed_df) * (min_per / 100))
        max_count = math.ceil(len(lemmed_df) * (max_per / 100))

        all_words = []
        ohe_dict = {}

        for col in input_df.columns:
            lemmed_df[col] = lemmed_df[col].apply(self.clean_text)

            lemmed_df[col] = lemmed_df[col].apply(
                lambda x: self.lemmatize_text(x, mystem, stop_words)
            )

            col_words = " ".join(lemmed_df[col]).split()
            all_words.extend(col_words)

            word_counts = Counter(col_words)
            filtered_words = {
                word: count
                for word, count in word_counts.items()
                if min_count < count < max_count
            }
            col_words_list = list(filtered_words.keys())

            col_ohe = pd.DataFrame(0, index=lemmed_df.index, columns=col_words_list)

            for word in col_words_list:
                col_ohe[word] = (
                    lemmed_df[col].str.contains(rf"\b{word}\b", regex=True).astype(int)
                )

            ohe_dict[col] = col_ohe

        return lemmed_df, ohe_dict
